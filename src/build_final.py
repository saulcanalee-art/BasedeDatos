import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
OUT="/sessions/adoring-peaceful-wright/mnt/outputs/Base_Datos_Investigadores.xlsx"
d=json.load(open("/sessions/adoring-peaceful-wright/mnt/outputs/resume_data.json"))
rows=d["rows"]; store=d["store"]; res2=json.load(open("/sessions/adoring-peaceful-wright/mnt/outputs/res2.json")); qres=json.load(open("/sessions/adoring-peaceful-wright/mnt/outputs/qres.json"))
name_by={}; 
for cvu,nm in rows: name_by.setdefault(cvu,nm)
pilotq={25712:[5,0,3,1,7],93301:[15,5,3,2,3],926347:[8,2,0,0,0],165173:[16,16,5,0,57],20585:[149,43,1,0,55],219087:[25,12,8,0,6],30281:[16,22,11,0,18],176679:[3,7,7,2,21],216216:[7,9,4,0,6]}

NAVY="1F4E78"; hf=PatternFill("solid",fgColor=NAVY); hfont=Font(name="Arial",bold=True,color="FFFFFF",size=10)
cf=Font(name="Arial",size=10); ctr=Alignment("center","center",wrap_text=True); lft=Alignment("left","center",wrap_text=True)
thin=Side(style="thin",color="D9D9D9"); bd=Border(thin,thin,thin,thin)
f_ok=PatternFill("solid",fgColor="E2EFDA"); f_rev=PatternFill("solid",fgColor="FFF2CC"); f_no=PatternFill("solid",fgColor="FCE4D6"); f_pend=PatternFill("solid",fgColor="F2F2F2")
wb=openpyxl.Workbook(); ws=wb.active; ws.title="Base de Datos"
H=["CVU","Nombre Investigador","ORCID","URL ORCID","Fuente ORCID","Total Publicaciones","Total Citas","Indice H","Indice i10","Citas Promedio 2yr","Citas por Pub","Articulos de Revista","Preprints/Editoriales","Capitulos de Libro","Pub Q1","Pub Q2","Pub Q3","Pub Q4","Sin Cuartil","Revistas Principales","Estado del registro","Notas"]
for c,h in enumerate(H,1):
    x=ws.cell(1,c,h); x.fill=hf; x.font=hfont; x.alignment=ctr; x.border=bd
r=2
dashdata=[]
for cvu,nombre in rows:
    v=store.get(str(cvu),{})
    oaid=v.get("oaid"); orcid=v.get("orcid")
    url=f"https://orcid.org/{orcid}" if orcid else ""
    fuente=("OpenAlex + ORCID" if (orcid and oaid) else ("ORCID (registro)" if orcid else ("OpenAlex (sin ORCID)" if oaid else "-")))
    pubs,cites,h,i10,p2=v.get("pubs"),v.get("cites"),v.get("h"),v.get("i10"),v.get("p2")
    cpp=round(cites/pubs,1) if (isinstance(cites,int) and isinstance(pubs,int) and pubs) else ("" if pubs is None else 0)
    estado=v.get("estado","Pendiente")
    r2=res2.get(str(cvu))
    if r2:
        art,pe,cap,rev=r2[0],r2[1],r2[2],r2[3]
    elif oaid:
        art=pe=cap=""; rev=""
    else:
        art=pe=cap=""; rev=""
    qq=qres.get(str(cvu))
    if qq: q1,q2,q3,q4,sq=qq
    else: q1=q2=q3=q4=sq=""
    notas = "" if oaid else "Sin perfil en OpenAlex; revisar en WoS/Scopus"
    if str(cvu) in ("854945",): notas="Posible persona distinta (revistas de oftalmologia); revisar identidad."
    vals=[cvu,nombre,orcid or "",url,fuente,pubs,cites,h,i10,p2,cpp,art,pe,cap,q1,q2,q3,q4,sq,rev,estado,notas]
    for c,val in enumerate(vals,1):
        x=ws.cell(r,c,val); x.font=cf; x.border=bd
        x.alignment=lft if c in (2,4,5,20,21,22) else ctr
    est=ws.cell(r,21)
    if estado.startswith("No encontrado"): est.fill=f_no
    elif estado.startswith("Pendiente"): est.fill=f_pend
    elif any(k in estado for k in("revisar","BUAP","confusion","Revisar")): est.fill=f_rev
    elif estado.startswith("Verificado"): est.fill=f_ok
    r+=1
W=[9,32,20,33,24,11,10,7,8,11,9,11,12,12,9,9,9,9,9,42,26,42]
for i,w in enumerate(W,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="C2"; ws.row_dimensions[1].height=30; ws.sheet_view.showGridLines=False

# ---- dash_data (uniques) ----
seen=set()
for cvu,nm in rows:
    if cvu in seen: continue
    seen.add(cvu)
    v=store.get(str(cvu),{}); r2=res2.get(str(cvu))
    rec={"cvu":cvu,"name":nm,"orcid":v.get("orcid"),"pubs":v.get("pubs"),"cites":v.get("cites"),"h":v.get("h"),"i10":v.get("i10"),"p2":v.get("p2"),"estado":v.get("estado","")}
    rec["cpp"]=round(rec["cites"]/rec["pubs"],1) if (isinstance(rec["cites"],int) and rec["pubs"]) else None
    if r2:
        rec["art"],rec["pe"],rec["cap"]=r2[0],r2[1],r2[2]
        rec["rev"]=[x for x in r2[3].split("; ") if x]
    if str(cvu) in qres: rec["q"]=qres[str(cvu)]
    dashdata.append(rec)
dashdata.sort(key=lambda r:-(r["cites"] or -1))
json.dump(dashdata,open("/sessions/adoring-peaceful-wright/mnt/outputs/dash_data.json","w"),ensure_ascii=False)

# ---- metrics + Datos sheet + Dashboard + Metodologia (igual que antes) ----
uni=[(int(c),name_by.get(int(c),""),v["pubs"],v["cites"],v["h"],v["i10"],v["p2"]) for c,v in store.items() if v.get("pubs") is not None]
tot=len(store); conperfil=sum(1 for v in store.values() if v.get("oaid")); conorcid=sum(1 for v in store.values() if v.get("orcid")); sinperfil=tot-conperfil
sumpubs=sum(x[2] for x in uni); sumcites=sum(x[3] for x in uni)
import statistics as st
havg=round(st.mean(x[4] for x in uni),1); hmax=max(x[4] for x in uni)
dz=wb.create_sheet("Datos")
def sn(n):p=n.split();return (p[0][0]+". "+" ".join(p[1:3])) if len(p)>2 else n
topC=sorted(uni,key=lambda x:-x[3])[:12]; topH=sorted(uni,key=lambda x:-x[4])[:12]
dz["A1"]="Inv";dz["B1"]="Citas"
for i,x in enumerate(topC,2): dz.cell(i,1,sn(x[1]));dz.cell(i,2,x[3])
dz["D1"]="Inv";dz["E1"]="H"
for i,x in enumerate(topH,2): dz.cell(i,4,sn(x[1]));dz.cell(i,5,x[4])
hb=[0]*6
for x in uni:
    hh=x[4]; hb[0 if hh<=5 else 1 if hh<=10 else 2 if hh<=15 else 3 if hh<=20 else 4 if hh<=30 else 5]+=1
dz["G1"]="Rango H";dz["H1"]="N"
for i,lab in enumerate(["0-5","6-10","11-15","16-20","21-30","31+"],2): dz.cell(i,7,lab);dz.cell(i,8,hb[i-2])
dz["J1"]="ORCID";dz["K1"]="N";dz["J2"]="Con ORCID";dz["K2"]=conorcid;dz["J3"]="Perfil s/ORCID";dz["K3"]=conperfil-conorcid;dz["J4"]="Sin perfil";dz["K4"]=sinperfil
# quartil agregado
totq=[0,0,0,0,0]
for v in qres.values():
    for i in range(5): totq[i]+=v[i]
labsq=["Q1","Q2","Q3","Q4","Sin cuartil"]
dz["M1"]="Cuartil";dz["N1"]="Publicaciones"
for i,lab in enumerate(labsq,2): dz.cell(i,13,lab);dz.cell(i,14,totq[i-2])
dz.sheet_state="hidden"
db=wb.create_sheet("Dashboard"); db.sheet_view.showGridLines=False; db.column_dimensions["A"].width=2
for col in "BCDEFGHIJKLMNOPQR": db.column_dimensions[col].width=11
db.merge_cells("B2:R2"); t=db["B2"]; t.value="PANEL DE INDICADORES — Investigadores UANL"; t.font=Font(name="Arial",bold=True,size=18,color=NAVY); t.alignment=Alignment("center","center")
db.merge_cells("B3:R3"); s=db["B3"]; s.value="Fuente: OpenAlex (jul-2026) · Cuartil SJR/Scimago (reemplazable por JCR)"; s.font=Font(name="Arial",size=9,italic=True,color="808080"); s.alignment=Alignment("center","center")
cards=[("Investigadores",tot,"1F4E78"),("Con perfil",conperfil,"2E7D32"),("Con ORCID",conorcid,"00838F"),("Sin perfil",sinperfil,"C62828"),("Publicaciones",f"{sumpubs:,}","5E35B1"),("Citas totales",f"{sumcites:,}","AD1457"),("H promedio",havg,"EF6C00"),("H maximo",hmax,"37474F")]
col=2;row=5
for lab,val,color in cards:
    c1=get_column_letter(col);c2=get_column_letter(col+3); db.merge_cells(f"{c1}{row}:{c2}{row+2}")
    cell=db[f"{c1}{row}"]; cell.value=f"{val}\n{lab}"; cell.fill=PatternFill("solid",fgColor=color); cell.font=Font(name="Arial",bold=True,size=15,color="FFFFFF"); cell.alignment=Alignment("center","center",wrap_text=True)
    col+=4
    if col>14: col=2;row+=4
def barc(title,minc,maxc,catcol,anchor,color):
    ch=BarChart();ch.type="bar";ch.title=title;ch.legend=None;ch.height=7.5;ch.width=13
    ch.add_data(Reference(dz,min_col=minc,min_row=1,max_row=maxc),titles_from_data=True); ch.set_categories(Reference(dz,min_col=catcol,min_row=2,max_row=maxc))
    try: ch.series[0].graphicalProperties.solidFill=color
    except: pass
    db.add_chart(ch,anchor)
barc("Top 12 por Citas",2,13,1,"B14","AD1457"); barc("Top 12 por Indice H",5,13,4,"J14","1F4E78")
colh=BarChart();colh.type="col";colh.title="Distribucion de Indice H";colh.legend=None;colh.height=7.5;colh.width=13
colh.add_data(Reference(dz,min_col=8,min_row=1,max_row=7),titles_from_data=True);colh.set_categories(Reference(dz,min_col=7,min_row=2,max_row=7))
try: colh.series[0].graphicalProperties.solidFill="EF6C00"
except: pass
db.add_chart(colh,"B31")
pie=PieChart();pie.title="Cobertura ORCID";pie.height=7.5;pie.width=10
pie.add_data(Reference(dz,min_col=11,min_row=1,max_row=4),titles_from_data=True);pie.set_categories(Reference(dz,min_col=10,min_row=2,max_row=4))
db.add_chart(pie,"J31")
cq=BarChart();cq.type="col";cq.title="Publicaciones por cuartil (SJR)";cq.legend=None;cq.height=7.5;cq.width=13
cq.add_data(Reference(dz,min_col=14,min_row=1,max_row=6),titles_from_data=True);cq.set_categories(Reference(dz,min_col=13,min_row=2,max_row=6))
try: cq.series[0].graphicalProperties.solidFill="2E7D32"
except: pass
db.add_chart(cq,"B48")
mt=wb.create_sheet("Metodologia");mt.column_dimensions["A"].width=110
N=[("BASE DE DATOS DE INVESTIGADORES UANL",True),("",False),
("93 investigadores (85 con perfil OpenAlex; 70 con ORCID; 8 sin perfil).",False),
("COMPLETO: ORCID, publicaciones, citas, H, i10, citas 2yr, citas/pub, tipo de trabajo y Revistas Principales.",False),
("CUARTILES (Pub Q1-Q4): SJR de Scimago (Best Quartile, ano vigente). JCR no disponible: la cuenta WoS no tiene suscripcion a JCR.",False),
("Se pueden reemplazar por JCR oficial si se obtiene acceso (red/VPN UANL). 'Sin Cuartil' = repositorios, congresos, preprints y revistas no indexadas en Scimago.",False),
("Fuentes: archivo original (CVU/Nombre) + OpenAlex (metricas, tipos, revistas). Desambiguacion por nombre completo+ORCID+afiliacion UANL.",False),
("Nota: los tipos no incluyen 'conference-papers', por eso pueden no sumar el total de publicaciones.",False),
("Revisar identidad: 854945 (revistas de oftalmologia, posible homonimo). Sin perfil: 169343,935315,120727,664945,479737,81182,219291,1144698.",False)]
for i,(txt,b) in enumerate(N,1):
    x=mt.cell(i,1,txt);x.font=Font(name="Arial",size=11,bold=b,color=NAVY if b else "000000");x.alignment=Alignment("left","center",wrap_text=True)
wb._sheets.sort(key=lambda sh:["Dashboard","Base de Datos","Metodologia","Datos"].index(sh.title))
wb.active=0
wb.save(OUT)
print("xlsx OK | dashdata:",len(dashdata),"| con revistas:",sum(1 for r in dashdata if r.get('rev')))
