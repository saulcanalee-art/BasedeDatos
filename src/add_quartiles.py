import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "/sessions/adoring-peaceful-wright/mnt/outputs/Base_Datos_Investigadores.xlsx"

# --- Mapa aproximado de cuartil SJR por revista (mejor categoria, ~2023/2024) ---
# SIN = repositorio/preprint/congreso/serie/no clasificable
Q = {}
def add(names,q):
    for n in names: Q[n.lower()] = q
add(["Mathematics","Journal of Process Control","Mathematics and Computers in Simulation",
     "Materials","Journal of Computational and Applied Mathematics",
     "Physica A Statistical Mechanics and its Applications","Advances in Computational Mathematics",
     "Applied Mathematics and Computation","Neurocomputing","ACS Applied Nano Materials",
     "Surfaces and Interfaces","Nanoscale","ACS Applied Energy Materials","Applied Surface Science",
     "Physical review. D/Physical review. D.","Physical Review D","Monthly Notices of the Royal Astronomical Society",
     "Astronomy and Astrophysics","Physical Review C","Journal of Cosmology and Astroparticle Physics",
     "Astroparticle Physics","Engineering Applications of Artificial Intelligence","Heliyon",
     "Computers & Mathematics with Applications","Journal of Cleaner Production",
     "Journal of the Franklin Institute","IEEE Transactions on Automatic Control","IEEE Transactions on Cybernetics",
     "International Journal of Systems Science","Automatica","IEEE Transactions on Fuzzy Systems",
     "IEEE Transactions on Industrial Electronics","International Journal of Robust and Nonlinear Control",
     "IEEE Transactions on Systems Man and Cybernetics Systems","Signal Processing","Information Sciences",
     "IEEE/ASME Transactions on Mechatronics","IEEE Transactions on Automation Science and Engineering",
     "Computers & Operations Research","Expert Systems with Applications","Computers & Industrial Engineering",
     "International Transactions in Operational Research","Annals of Operations Research","Omega",
     "Applied Soft Computing","Swarm and Evolutionary Computation","Socio-Economic Planning Sciences",
     "Optics & Laser Technology","Journal of Lightwave Technology","IEEE Access","Computer Networks",
     "Optics Letters","Optics Express","Journal of Quality Technology","IEEE Journal on Selected Areas in Communications",
     "IEEE Communications Surveys & Tutorials","Optica","Journal of Network and Computer Applications",
     "IEEE Transactions on Industrial Informatics","Automation and Remote Control","Systems & Control Letters",
     "IEEE/CAA Journal of Automatica Sinica","Chaos Solitons & Fractals","Photonics Research",
     "Journal of Optical Communications and Networking","Nonlinear Analysis Hybrid Systems"], "Q1")
add(["ACS Omega","New Journal of Chemistry","Universe","European Physical Journal A","Symmetry","Sensors",
     "Earth Science Informatics","Frontiers in Genetics","Computing","Journal of Statistical Mechanics Theory and Experiment",
     "The European Physical Journal Special Topics","Journal of Atmospheric and Solar-Terrestrial Physics",
     "Applied Sciences","Physica Scripta","IET Control Theory and Applications","Applied Optics","Optics Communications",
     "Optical Fiber Technology","Quality and Reliability Engineering International","Optimization",
     "Circuits Systems and Signal Processing","Asian Journal of Control","Transactions of the Institute of Measurement and Control",
     "International Journal of Adaptive Control and Signal Processing","International Journal of General Systems",
     "International Journal of Control","Optimal Control Applications and Methods","Complexity","Axioms",
     "Optical and Quantum Electronics","Wireless Personal Communications","Electronics Letters","Physical Communication",
     "IEEE Communications Letters","IEEE Photonics Technology Letters","Journal of Statistical Computation and Simulation",
     "Communications in Statistics - Simulation and Computation","Computational Statistics","Applied Economics",
     "The International Journal of Advanced Manufacturing Technology","RAIRO - Operations Research",
     "Computational Optimization and Applications","Operational Research","Networks","Optimization Letters",
     "Transportmetrica A Transport Science","Engineering Computations","Memetic Computing","PLoS ONE",
     "Journal of High Speed Networks","Annals of Telecommunications","EURASIP Journal on Wireless Communications and Networking",
     "The International Journal of Logistics Management","Progress In Electromagnetics Research B"], "Q2")
add(["Mathematical Problems in Engineering","DYNA","Frontiers in Applied Mathematics and Statistics",
     "Physics of Particles and Nuclei","Astronomische Nachrichten","Physics Education","INTELIGENCIA ARTIFICIAL",
     "Laser Physics","Photonic Network Communications","Optics","International Journal of Antennas and Propagation",
     "Journal of Sensors","Microwave and Optical Technology Letters","Journal of Applied Research and Technology",
     "Gerontechnology","Kybernetika","Journal of Mathematics","Journal of Applied Mathematics and Physics",
     "International Journal of Prognostics and Health Management","Recent Patents on Electrical Engineering"], "Q3")
add(["Revista Mexicana de Fisica","Revista Mexicana de Física","Revista Mexicana de Física E",
     "Tecnologia y Ciencias del Agua","Tecnología y Ciencias del Agua","Acta Universitaria",
     "Ingenieria Investigacion y Tecnologia","Ingeniería Investigación y Tecnología","Computacion y Sistemas",
     "Computación y Sistemas","Research in Computing Science"], "Q4")

def cls(name):
    n=name.lower()
    if n in Q: return Q[n]
    # repositorios / preprints / congresos / series -> SIN
    sin_kw=["arxiv","research square","preprints.org","ssrn","repository","repositório","hal ",
            "doaj","dialnet","pubmed","figshare","conference","proceedings","lecture notes",
            "pos(","icrc","aip conference","epj web","eBooks".lower(),"ebooks","referencia","conicet",
            "database","world scientific","storage and retrieval","acta physica polonica b proceedings",
            "suplemento","cern","victoria university","bielefeld","zagreb","bologna","academy",
            "communications in computer and information science","studies in systems","spie",
            "optical fiber communication conference","european conference on optical","optical network design",
            "latin america optics","optical fiber sensors","dataset","paratext","dyna energia","dyna management",
            "eai/springer","cleaner logistics","aims environmental","international journal of combinatorial"]
    if any(k in n for k in sin_kw): return "SIN"
    return "NC"   # journal real no mapeado

# --- distribucion de fuentes por autor (OpenAlex group_by source) name:count ---
authors = {
25712:{"Mathematics":2,"DYNA":1,"Mathematical Problems in Engineering":1,"Materials":1,
       "Mathematics and Computers in Simulation":1,"Tecnología y Ciencias del Agua":1,
       "INTELIGENCIA ARTIFICIAL":1,"Journal of Process Control":1,
       "IFAC Proceedings Volumes":2,"Lecture notes in networks and systems":1,"DYNA ENERGIA Y SOSTENIBILIDAD":1,
       "DOAJ":1,"Dialnet":1,"Preprints.org":1},
93301:{"Journal of Computational and Applied Mathematics":3,"Physica A Statistical Mechanics and its Applications":3,
       "Advances in Computational Mathematics":3,"Earth Science Informatics":2,
       "Frontiers in Applied Mathematics and Statistics":2,"Applied Mathematics and Computation":2,
       "Journal of Statistical Mechanics Theory and Experiment":1,"Mathematical Problems in Engineering":1,
       "Revista Mexicana de Física":1,"Revista Mexicana de Física E":1,"Frontiers in Genetics":1,"Heliyon":1,
       "Computing":1,"Computers & Mathematics with Applications":1,"Neurocomputing":1,"Journal of Cleaner Production":1,
       "Research Square":3},
926347:{"ACS Applied Nano Materials":3,"Surfaces and Interfaces":2,"Nanoscale":1,"New Journal of Chemistry":1,
        "ACS Applied Energy Materials":1,"ACS Omega":1,"Applied Surface Science":1},
165173:{"Physical review. D/Physical review. D.":6,"Universe":4,"Physics of Particles and Nuclei":3,
        "European Physical Journal A":3,"Symmetry":3,"Sensors":2,"Monthly Notices of the Royal Astronomical Society":2,
        "Astronomy and Astrophysics":2,"Physical Review C":2,"Journal of Cosmology and Astroparticle Physics":1,
        "Physical Review D":1,"The European Physical Journal Special Topics":1,"Astronomische Nachrichten":1,
        "Astroparticle Physics":1,"Journal of Atmospheric and Solar-Terrestrial Physics":1,"Applied Sciences":1,
        "Physics Education":1,"Physica Scripta":1,"Engineering Applications of Artificial Intelligence":1,
        "arXiv (Cornell University)":33,"PoS ICRC2021":8,"Journal of Physics Conference Series":4,"PoS ICRC2019":3,
        "AIP conference proceedings":2,"Acta Physica Polonica B Proceedings Supplement":2,"Preprints.org":2,
        "EPJ Web of Conferences":1,"SSRN Electronic Journal":1,"Suplemento":1},
20585:{"Journal of the Franklin Institute":29,"International Journal of Systems Science":15,
       "IEEE Transactions on Automatic Control":12,"IEEE Transactions on Cybernetics":12,
       "IEEE Transactions on Systems Man and Cybernetics Systems":12,"IET Control Theory and Applications":10,
       "IEEE Transactions on Industrial Electronics":10,"International Journal of Robust and Nonlinear Control":9,
       "IEEE Transactions on Fuzzy Systems":7,"IEEE/ASME Transactions on Mechatronics":7,
       "Transactions of the Institute of Measurement and Control":7,"Asian Journal of Control":6,
       "Circuits Systems and Signal Processing":6,"Signal Processing":5,"Information Sciences":5,"Automatica":5,
       "International Journal of Adaptive Control and Signal Processing":5,"International Journal of General Systems":4,
       "IEEE/CAA Journal of Automatica Sinica":4,"IEEE Transactions on Automation Science and Engineering":4,
       "International Journal of Stochastic Analysis":4,"IEEE Access":3,"Systems & Control Letters":3,
       "International Journal of Control":3,"Automation and Remote Control":1,"Neurocomputing":2,
       "Optimal Control Applications and Methods":2,"Chaos Solitons & Fractals":1,"Nonlinear Analysis Hybrid Systems":1,
       "Applied Soft Computing":1,"Applied Mathematics and Computation":1,"Mathematical Problems in Engineering":1,
       "IFAC Proceedings Volumes":27,"Lecture notes in control and information sciences":9,"IFAC-PapersOnLine":4,
       "The Journal of Urology":4,"arXiv (Cornell University)":4,"PubMed":3},
219087:{"Computers & Operations Research":5,"Mathematical Problems in Engineering":5,"Expert Systems with Applications":3,
        "Computers & Industrial Engineering":3,"International Transactions in Operational Research":3,
        "Applied Soft Computing":2,"Applied Mathematics and Computation":2,"Annals of Operations Research":2,
        "Journal of Cleaner Production":2,"Memetic Computing":1,"Optimization Letters":1,"DYNA":1,"Operational Research":1,
        "Socio-Economic Planning Sciences":1,"Networks":1,"PLoS ONE":1,"Complexity":1,"Swarm and Evolutionary Computation":1,
        "Journal of Mathematics":1,"Computational Optimization and Applications":1,"Axioms":1,
        "Transportmetrica A Transport Science":1,"Omega":1,"Journal of Applied Research and Technology":1,
        "Engineering Computations":1,"The International Journal of Logistics Management":1,"RAIRO - Operations Research":1,
        "Lecture notes in computer science":2,"arXiv (Cornell University)":3,"SSRN Electronic Journal":1},
30281:{"Photonic Network Communications":8,"Optics & Laser Technology":3,"IEEE Access":3,"Applied Sciences":3,
       "Applied Optics":3,"Computer Networks":3,"Sensors":2,"Physical Communication":2,"Electronics Letters":2,
       "Mathematical Problems in Engineering":2,"Journal of Lightwave Technology":2,"Wireless Personal Communications":2,
       "IEEE Transactions on Broadcasting":1,"Optical and Quantum Electronics":1,"Journal of High Speed Networks":1,
       "IEEE Communications Letters":1,"International Journal of Antennas and Propagation":1,"Annals of Telecommunications":1,
       "Optics Express":1,"Journal of Optical Communications and Networking":1,"IEEE Communications Surveys & Tutorials":1,
       "IEEE Photonics Technology Letters":1,"Journal of Network and Computer Applications":1,
       "Progress In Electromagnetics Research B":1,"Optical Fiber Technology":1,
       "EURASIP Journal on Wireless Communications and Networking":1,"IEEE Journal on Selected Areas in Communications":1,
       "Proceedings of SPIE":6,"HAL":5,"IGI Global eBooks":3,"Elsevier eBooks":2,"Lecture notes in computer science":1},
176679:{"Sensors":4,"Laser Physics":4,"Optics Communications":2,"Optics Letters":1,"Optics Express":1,
        "IEEE Sensors Journal":1,"Energies":1,"Acta Universitaria":1,"Catalysts":1,"Photonics Research":1,
        "Advanced Photonics Research":1,"Optics":1,"Microwave and Optical Technology Letters":1,
        "Optical Fiber Technology":1,"Journal of Sensors":1,"Research in Computing Science":1,
        "Proceedings of SPIE":13,"Latin America Optics and Photonics Conference":3,"Optical Fiber Sensors":1},
216216:{"Quality and Reliability Engineering International":2,"Computers & Industrial Engineering":2,
        "Expert Systems with Applications":1,"Operational Research":1,
        "Communications in Statistics - Simulation and Computation":1,"Optimization":1,"Journal of Quality Technology":1,
        "Mathematical Problems in Engineering":1,"Journal of Statistical Computation and Simulation":1,
        "Applied Economics":1,"Astronomy and Astrophysics":1,"Journal of Applied Mathematics and Physics":1,
        "International Journal of Prognostics and Health Management":1,"Mathematics":1,"Applied Mathematics and Computation":1,
        "Gerontechnology":1,"The International Journal of Advanced Manufacturing Technology":1,"Computational Statistics":1,
        "arXiv (Cornell University)":3,"Communications in computer and information science":1,"PubMed":1,"ArXiv.org":1}
}

# calcular
res={}
for cvu,src in authors.items():
    c={"Q1":0,"Q2":0,"Q3":0,"Q4":0,"SIN":0,"NC":0}
    for name,n in src.items():
        c[cls(name)] += n
    clasif=c["Q1"]+c["Q2"]+c["Q3"]+c["Q4"]
    tot_journal=clasif+c["NC"]
    res[cvu]=(c,clasif,tot_journal)
    print(cvu,c,"cobertura %.0f%%"%(100*clasif/max(1,tot_journal)))

# --- escribir en el workbook ---
wb=openpyxl.load_workbook(OUT)
ws=wb["Base de Datos"]
cell_font=Font(name="Arial",size=10)
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
# encabezado: renombrar Estado -> Estado del registro
ws.cell(1,20).value="Estado del registro"
# localizar filas por CVU y rellenar Q (cols 14-18) y anexar cobertura a Notas(21)
for r in range(2,ws.max_row+1):
    cvu=ws.cell(r,1).value
    if cvu in res:
        c,clasif,totj=res[cvu]
        ws.cell(r,14,c["Q1"]); ws.cell(r,15,c["Q2"]); ws.cell(r,16,c["Q3"]); ws.cell(r,17,c["Q4"])
        ws.cell(r,18,c["SIN"]+c["NC"])
        for col in (14,15,16,17,18):
            ws.cell(r,col).font=cell_font; ws.cell(r,col).alignment=center
        cov="Cuartil SJR APROX (nivel revista). Cobertura %d/%d art. de revista clasificados; %d en repos/congreso/preprint (Sin cuartil)."%(clasif,totj,c["SIN"])
        old=ws.cell(r,21).value or ""
        ws.cell(r,21).value=(old+" | " if old else "")+cov

wb.save(OUT)
print("OK actualizado")
