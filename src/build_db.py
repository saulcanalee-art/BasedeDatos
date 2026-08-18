import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/sessions/adoring-peaceful-wright/mnt/uploads/LISTADO INVESTIGADORES (1).xlsx"
OUT = "/sessions/adoring-peaceful-wright/mnt/outputs/Base_Datos_Investigadores.xlsx"

# --- leer lista original (CVU, Nombre) ---
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_src = wb_src["Hoja1"]
researchers = []
for row in ws_src.iter_rows(min_row=4, values_only=True):
    cvu, nombre = row[1], row[2]
    if cvu is None and nombre is None:
        continue
    researchers.append((cvu, str(nombre).strip() if nombre else ""))

# --- datos del piloto (OpenAlex), keyed por CVU ---
# campos: orcid, pubs, cites, h, i10, prom2yr, art, prep_ed, cap, revistas, notas, estado
pilot = {
 25712: ("0000-0002-9079-2771",28,134,6,2,3.33,13,1,0,
   "Mathematics; Mathematical Problems in Engineering; Journal of Process Control; Mathematics and Computers in Simulation; Materials",
   "14 conference-papers no contados en columnas de tipo; cuartiles pendientes","Verificado"),
 93301: ("0000-0001-9422-5259",28,202,6,3,1.67,24,4,0,
   "Journal of Computational and Applied Mathematics; Physica A; Advances in Computational Mathematics; Applied Mathematics and Computation; Neurocomputing",
   "Cuartiles pendientes","Verificado"),
 926347: ("0009-0002-5608-3119",10,72,6,2,4.60,10,0,0,
   "ACS Applied Nano Materials; Surfaces and Interfaces; Nanoscale; New Journal of Chemistry; Applied Surface Science",
   "ORCID nuevo (formato 0009-); cuartiles pendientes","Verificado"),
 165173: ("0000-0003-1874-8116",115,1785,19,39,5.67,47,36,1,
   "Physical Review D; Universe; Physics of Particles and Nuclei; European Physical Journal A; MNRAS; Astronomy & Astrophysics",
   "27 conference-papers; 33 preprints en arXiv; cuartiles pendientes","Verificado"),
 169343: (None,None,None,None,None,None,None,None,None,"","Sin perfil localizable en OpenAlex por nombre","No encontrado - revisar en WoS/Scopus"),
 20585: ("0000-0002-7274-4303",470,8745,48,152,2.16,242,17,9,
   "Journal of the Franklin Institute; IEEE Trans. Automatic Control; IEEE Trans. Cybernetics; International Journal of Systems Science; Automatica; IEEE Trans. Fuzzy Systems",
   "200 conference-papers; nombre indexado como 'Michael/M.V. Basin'; cuartiles pendientes","Verificado"),
 219087: ("0000-0001-6519-7690",59,989,16,23,1.33,50,5,0,
   "Computers & Operations Research; Expert Systems with Applications; Computers & Industrial Engineering; International Transactions in Operational Research; Annals of Operations Research; Omega",
   "Cuartiles pendientes","Verificado"),
 30281: ("0000-0001-5208-5745",127,1210,16,34,2.89,60,1,5,
   "Photonic Network Communications; Optics & Laser Technology; Applied Optics; Journal of Lightwave Technology; IEEE Access; Computer Networks",
   "60 conference-papers (SPIE/OFC); afiliacion principal Tec de Monterrey + UANL; cuartiles pendientes","Verificado (afil. ITESM/UANL)"),
 176679: ("0000-0002-0505-832X",54,753,12,16,0.00,23,0,0,
   "Laser Physics; Optics Communications; Optics Letters; Optics Express; Sensors; Optical Fiber Technology",
   "31 conference-papers (SPIE); temas de fibra optica confirman identidad; indexado como 'A. Castillo-Guzman'; cuartiles pendientes","Verificado (revisar)"),
 216216: ("0000-0003-4389-7682",28,229,8,8,1.00,21,4,0,
   "Quality and Reliability Engineering International; Computers & Industrial Engineering; Journal of Quality Technology; Expert Systems with Applications; Optimization",
   "Cuartiles pendientes","Verificado"),
}

headers = ["CVU","Nombre Investigador","ORCID","URL ORCID","Fuente ORCID",
 "Total Publicaciones","Total Citas","Indice H","Indice i10","Citas Promedio 2yr",
 "Articulos de Revista","Preprints/Editoriales","Capitulos de Libro",
 "Pub Q1","Pub Q2","Pub Q3","Pub Q4","Sin Cuartil","Revistas Principales","Estado","Notas"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Base de Datos"

# estilos
hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
cell_font = Font(name="Arial", size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
fill_ok = PatternFill("solid", fgColor="E2EFDA")     # verde claro verificado
fill_rev = PatternFill("solid", fgColor="FFF2CC")    # amarillo revisar
fill_no = PatternFill("solid", fgColor="FCE4D6")     # naranja no encontrado
fill_pend = PatternFill("solid", fgColor="F2F2F2")   # gris pendiente

for c,h in enumerate(headers,1):
    cell = ws.cell(1,c,h)
    cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=border

r = 2
for cvu, nombre in researchers:
    d = pilot.get(cvu)
    orcid = d[0] if d else None
    url = f"https://orcid.org/{orcid}" if orcid else None
    fuente = "OpenAlex -> ORCID verificado" if orcid else ("OpenAlex (sin match)" if d else "Pendiente")
    if d:
        pubs,cites,h,i10,p2 = d[1],d[2],d[3],d[4],d[5]
        art,pe,cap = d[6],d[7],d[8]
        rev,notas,estado = d[9],d[10],d[11]
        q = "Pendiente" if d[1] is not None else ""
    else:
        pubs=cites=h=i10=p2=art=pe=cap=None
        rev=""; notas=""; estado="Pendiente (fase siguiente)"; q=""
    vals=[cvu,nombre,orcid or "",url or "",fuente,pubs,cites,h,i10,p2,art,pe,cap,
          q,q,q,q,q,rev,estado,notas]
    for c,v in enumerate(vals,1):
        cell=ws.cell(r,c,v); cell.font=cell_font; cell.border=border
        cell.alignment=left if c in (2,4,5,19,20,21) else center
    # colorear Estado
    est_cell=ws.cell(r,20)
    if estado.startswith("Verificado (revisar"):
        est_cell.fill=fill_rev
    elif estado.startswith("Verificado"):
        est_cell.fill=fill_ok
    elif estado.startswith("No encontrado"):
        est_cell.fill=fill_no
    elif estado.startswith("Pendiente"):
        est_cell.fill=fill_pend
    else:
        est_cell.fill=fill_rev
    r+=1

# anchos
widths=[9,32,20,34,24,11,10,8,9,12,11,12,12,7,7,7,7,8,46,26,40]
for i,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="C2"
ws.row_dimensions[1].height=30

# --- hoja Metodologia ---
ws2=wb.create_sheet("Metodologia y Fuentes")
notes=[
 ("BASE DE DATOS DE INVESTIGADORES - Metodologia",True),
 ("",False),
 ("Fecha de elaboracion: 28-jul-2026",False),
 ("Total de registros en el listado original: 102 filas (~95 investigadores unicos; hay duplicados).",False),
 ("Fase actual: PILOTO de 10 investigadores enriquecidos para validacion.",False),
 ("",False),
 ("FUENTES DE DATOS",True),
 ("1. CVU y Nombre: del archivo original proporcionado por el usuario.",False),
 ("2. ORCID, Total Publicaciones, Total Citas, Indice H, Indice i10, Citas Promedio 2yr,",False),
 ("   tipos de trabajo y Revistas Principales: OpenAlex (api.openalex.org), snapshot jul-2026.",False),
 ("3. Cuartiles (Q1-Q4): PENDIENTES. Se completaran en fase posterior con Scimago/JCR (Web of Science)",False),
 ("   y se confirmaran contigo antes de fijarlos, segun lo acordado.",False),
 ("",False),
 ("NOTAS METODOLOGICAS IMPORTANTES",True),
 ("- ORCID NO contiene metricas bibliometricas (H, citas, cuartiles). Por eso el grueso de",False),
 ("  las metricas proviene de OpenAlex, no de orcid.org.",False),
 ("- Desambiguacion: se filtro por institucion (UANL, OpenAlex I169046204) y se emparejo por",False),
 ("  NOMBRE COMPLETO, no solo apellido. En varios casos el registro con mas trabajos correspondia",False),
 ("  a un homonimo distinto (p.ej. 'Efrain Alcorta' vs 'Maria Aracelia Alcorta').",False),
 ("- OpenAlex a veces fragmenta a una persona en varios IDs; se eligio el registro dominante",False),
 ("  con ORCID verificado. Las metricas son aproximadas y pueden diferir de WoS/Scopus.",False),
 ("- Columnas de tipo (Articulos/Preprints-Editoriales/Capitulos) NO incluyen 'conference-papers',",False),
 ("  que en ingenieria/fisica pueden ser numerosos; por eso pueden no sumar el Total Publicaciones.",False),
 ("  El detalle de congresos se anota en la columna Notas.",False),
 ("",False),
 ("LEYENDA DE 'ESTADO'",True),
 ("Verificado = match de alta confianza (nombre completo + ORCID + afiliacion).",False),
 ("Verificado (revisar) = match probable; conviene confirmar manualmente.",False),
 ("No encontrado = sin perfil localizable en OpenAlex; revisar en Web of Science/Scopus.",False),
 ("Pendiente (fase siguiente) = aun no procesado (fuera del piloto de 10).",False),
 ("",False),
 ("CASO SIN MATCH EN EL PILOTO",True),
 ("- Andres Alberto Aviles Alvarado (CVU 169343): no se localizo perfil por nombre en OpenAlex.",False),
 ("  Se recomienda buscarlo en Web of Science con tus credenciales o confirmar su nombre de publicacion.",False),
]
ws2.column_dimensions["A"].width=105
for i,(txt,bold) in enumerate(notes,1):
    c=ws2.cell(i,1,txt)
    c.font=Font(name="Arial",size=11,bold=bold,color="1F4E78" if bold else "000000")
    c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb.save(OUT)
print("Guardado:",OUT)
print("Filas de datos:",r-2)
