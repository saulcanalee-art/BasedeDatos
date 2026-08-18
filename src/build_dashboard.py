import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

OUT="/sessions/adoring-peaceful-wright/mnt/outputs/Base_Datos_Investigadores.xlsx"
d=json.load(open("/sessions/adoring-peaceful-wright/mnt/outputs/resume_data.json"))
rows=d["rows"]; store=d["store"]
name_by_cvu={}
for cvu,nm in rows:
    name_by_cvu.setdefault(cvu,nm)

pilot={
25712:(13,1,0,[5,0,3,1,7],"Mathematics; Math. Problems in Eng.; J. Process Control; Math and Computers in Simulation; Materials"),
93301:(24,4,0,[15,5,3,2,3],"J. Comp. & Applied Math.; Physica A; Adv. in Comp. Math.; Applied Math. & Comp.; Neurocomputing"),
926347:(10,0,0,[8,2,0,0,0],"ACS Applied Nano Materials; Surfaces and Interfaces; Nanoscale; New J. of Chemistry; Applied Surface Science"),
165173:(47,36,1,[16,16,5,0,57],"Physical Review D; Universe; Physics of Particles and Nuclei; Eur. Phys. J. A; MNRAS; A&A"),
20585:(242,17,9,[149,43,1,0,55],"J. Franklin Institute; IEEE T. Automatic Control; IEEE T. Cybernetics; Int. J. Systems Science; Automatica"),
219087:(50,5,0,[25,12,8,0,6],"Computers & Oper. Research; Expert Systems w/ Appl.; Computers & Ind. Eng.; ITOR; Annals of OR; Omega"),
30281:(60,1,5,[16,22,11,0,18],"Photonic Network Comm.; Optics & Laser Tech.; Applied Optics; J. Lightwave Tech.; IEEE Access"),
176679:(23,0,0,[3,7,7,2,21],"Laser Physics; Optics Communications; Optics Letters; Optics Express; Sensors; Optical Fiber Tech."),
216216:(21,4,0,[7,9,4,0,6],"Quality & Reliab. Eng. Int.; Computers & Ind. Eng.; J. Quality Tech.; Expert Systems w/ Appl.; Optimization"),
}

# ---------- Hoja Base de Datos ----------
wb=openpyxl.Workbook(); ws=wb.active; ws.title="Base de Datos"
H=["CVU","Nombre Investigador","ORCID","URL ORCID","Fuente ORCID","Total Publicaciones","Total Citas","Indice H","Indice i10","Citas Promedio 2yr","Citas por Pub","Articulos de Revista","Preprints/Editoriales","Capitulos de Libro","Pub Q1","Pub Q2","Pub Q3","Pub Q4","Sin Cuartil","Revistas Principales","Estado del registro","Notas"]
NAVY="1F4E78"; hf=PatternFill("solid",fgColor=NAVY); hfont=Font(name="Arial",bold=True,color="FFFFFF",size=10)
cf=Font(name="Arial",size=10); ctr=Alignment("center","center",wrap_text=True); lft=Alignment("left","center",wrap_text=True)
thin=Side(style="thin",color="D9D9D9"); bd=Border(thin,thin,thin,thin)
f_ok=PatternFill("solid",fgColor="E2EFDA"); f_rev=PatternFill("solid",fgColor="FFF2CC"); f_no=PatternFill("solid",fgColor="FCE4D6"); f_pend=PatternFill("solid",fgColor="F2F2F2")
for c,h in enumerate(H,1):
    x=ws.cell(1,c,h); x.fill=hf; x.font=hfont; x.alignment=ctr; x.border=bd
r=2
for cvu,nombre in rows:
    v=store.get(str(cvu),{})
    oaid=v.get("oaid"); orcid=v.get("orcid")
    url=f"https://orcid.org/{orcid}" if orcid else ""
    fuente="OpenAlex -> ORCID verificado" if orcid else ("OpenAlex (sin ORCID)" if oaid else "-")
    pubs,cites,h,i10,p2=v.get("pubs"),v.get("cites"),v.get("h"),v.get("i10"),v.get("p2")
    cpp=round(cites/pubs,1) if (isinstance(cites,int) and isinstance(pubs,int) and pubs) else ("" if pubs is None else 0)
    estado=v.get("estado","Pendiente")
    p=pilot.get(cvu)
    if p:
        art,pe,cap=p[0],p[1],p[2]; q1,q2,q3,q4,sq=p[3]; rev=p[4]; notas=""
    elif oaid:
        art=pe=cap="Pendiente"; q1=q2=q3=q4=sq="Pend. JCR"; rev="Pendiente"; notas="Tipo/revistas: tarea programada. Cuartil: JCR con WoS."
    else:
        art=pe=cap=""; q1=q2=q3=q4=sq=""; rev=""; notas="Sin perfil en OpenAlex; revisar en WoS/Scopus"
    vals=[cvu,nombre,orcid or "",url,fuente,pubs,cites,h,i10,p2,cpp,art,pe,cap,q1,q2,q3,q4,sq,rev,estado,notas]
    for c,val in enumerate(vals,1):
        x=ws.cell(r,c,val); x.font=cf; x.border=bd
        x.alignment=lft if c in (2,4,5,20,21,22) else ctr
    est=ws.cell(r,21)
    if estado.startswith("No encontrado"): est.fill=f_no
    elif estado.startswith("Pendiente"): est.fill=f_pend
    elif any(k in estado for k in("revisar","BUAP","confusion","Revisar")): est.fill=f_rev
    elif estado.startswith("Verificado"): est.fill=f_ok
    r+=1
W=[9,32,20,33,24,11,10,7,8,11,9,11,12,12,7,7,7,7,8,40,26,40]
for i,w in enumerate(W,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="C2"; ws.row_dimensions[1].height=30; ws.sheet_view.showGridLines=False

# ---------- métricas (únicos con perfil) ----------
uni=[]
for cvu,v in store.items():
    if v.get("pubs") is not None:
        uni.append((int(cvu),name_by_cvu.get(int(cvu),""),v["pubs"],v["cites"],v["h"],v["i10"],v["p2"]))
tot=len(store); conperfil=sum(1 for v in store.values() if v.get("oaid")); conorcid=sum(1 for v in store.values() if v.get("orcid"))
sinperfil=sum(1 for v in store.values() if not v.get("oaid"))
sumpubs=sum(x[2] for x in uni); sumcites=sum(x[3] for x in uni)
import statistics as st
havg=round(st.mean(x[4] for x in uni),1); hmax=max(x[4] for x in uni)
citavg=round(sumcites/len(uni)); pubavg=round(sumpubs/len(uni),1)

# ---------- Hoja Datos (fuente de graficas) ----------
dz=wb.create_sheet("Datos"); 
topC=sorted(uni,key=lambda x:-x[3])[:12]
topH=sorted(uni,key=lambda x:-x[4])[:12]
topP=sorted(uni,key=lambda x:-x[2])[:12]
def shortname(n):
    parts=n.split(); return (parts[0][0]+". "+" ".join(parts[1:3])) if len(parts)>2 else n
# tabla top citas
dz["A1"]="Investigador"; dz["B1"]="Citas"
for i,x in enumerate(topC,2): dz.cell(i,1,shortname(x[1])); dz.cell(i,2,x[3])
dz["D1"]="Investigador"; dz["E1"]="Indice H"
for i,x in enumerate(topH,2): dz.cell(i,4,shortname(x[1])); dz.cell(i,5,x[4])
dz["G1"]="Investigador"; dz["H1"]="Publicaciones"
for i,x in enumerate(topP,2): dz.cell(i,7,shortname(x[1])); dz.cell(i,8,x[2])
# distribucion H
hbins=[("0-5",0),("6-10",0),("11-15",0),("16-20",0),("21-30",0),("31+",0)]
def hbin(h):
    return 0 if h<=5 else 1 if h<=10 else 2 if h<=15 else 3 if h<=20 else 4 if h<=30 else 5
hb=[0]*6
for x in uni: hb[hbin(x[4])]+=1
dz["J1"]="Rango H"; dz["K1"]="Investigadores"
for i,(lab,_) in enumerate(hbins,2): dz.cell(i,10,lab); dz.cell(i,11,hb[i-2])
# distribucion citas
cbins=["0-49","50-199","200-499","500-999","1000-2499","2500+"]
def cbin(c): return 0 if c<50 else 1 if c<200 else 2 if c<500 else 3 if c<1000 else 4 if c<2500 else 5
cb=[0]*6
for x in uni: cb[cbin(x[3])]+=1
dz["M1"]="Rango Citas"; dz["N1"]="Investigadores"
for i,lab in enumerate(cbins,2): dz.cell(i,13,lab); dz.cell(i,14,cb[i-2])
# ORCID pie
dz["P1"]="ORCID"; dz["Q1"]="N"
dz["P2"]="Con ORCID"; dz["Q2"]=conorcid; dz["P3"]="Sin ORCID (perfil s/ORCID)"; dz["Q3"]=conperfil-conorcid; dz["P4"]="Sin perfil"; dz["Q4"]=sinperfil
# Estado pie
estcat={"Verificado":0,"Revisar":0,"No encontrado":0}
for v in store.values():
    e=v.get("estado","")
    if e.startswith("No encontrado"): estcat["No encontrado"]+=1
    elif any(k in e for k in("revisar","Revisar","BUAP","confusion")): estcat["Revisar"]+=1
    elif e.startswith("Verificado"): estcat["Verificado"]+=1
    else: estcat["Revisar"]+=1
dz["S1"]="Estado"; dz["T1"]="N"
for i,(k,val) in enumerate(estcat.items(),2): dz.cell(i,19,k); dz.cell(i,20,val)
dz.sheet_state="hidden"

# ---------- Hoja Dashboard ----------
db=wb.create_sheet("Dashboard"); db.sheet_view.showGridLines=False
db.column_dimensions["A"].width=2
for col in "BCDEFGHIJKLMNOPQR": db.column_dimensions[col].width=11
db.merge_cells("B2:R2"); t=db["B2"]; t.value="PANEL DE INDICADORES — Investigadores UANL"; t.font=Font(name="Arial",bold=True,size=18,color=NAVY); t.alignment=Alignment("center","center")
db.merge_cells("B3:R3"); s=db["B3"]; s.value="Fuente: OpenAlex (jul-2026) · Cuartiles JCR pendientes (Web of Science)"; s.font=Font(name="Arial",size=9,italic=True,color="808080"); s.alignment=Alignment("center","center")
cards=[("Investigadores",tot,"1F4E78"),("Con perfil",conperfil,"2E7D32"),("Con ORCID",conorcid,"00838F"),("Sin perfil",sinperfil,"C62828"),
("Publicaciones",f"{sumpubs:,}","5E35B1"),("Citas totales",f"{sumcites:,}","AD1457"),("H promedio",havg,"EF6C00"),("H maximo",hmax,"37474F")]
col=2; row=5
for i,(lab,val,color) in enumerate(cards):
    c1=get_column_letter(col); c2=get_column_letter(col+3)
    db.merge_cells(f"{c1}{row}:{c2}{row+2}")
    cell=db[f"{c1}{row}"]; cell.value=f"{val}\n{lab}"; cell.fill=PatternFill("solid",fgColor=color)
    cell.font=Font(name="Arial",bold=True,size=15,color="FFFFFF"); cell.alignment=Alignment("center","center",wrap_text=True)
    col+=4
    if col>14: col=2; row+=4

def barchart(title,minc,maxc,catcol,anchor,horizontal=True,color="1F4E78"):
    ch=BarChart(); ch.type="bar" if horizontal else "col"; ch.title=title; ch.legend=None; ch.height=7.5; ch.width=13
    data=Reference(dz,min_col=minc,min_row=1,max_row=maxc)
    cats=Reference(dz,min_col=catcol,min_row=2,max_row=maxc)
    ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
    ch.y_axis.majorGridlines=None
    try: ch.series[0].graphicalProperties.solidFill=color
    except: pass
    db.add_chart(ch,anchor)

# charts
barchart("Top 12 por Citas totales",2,13,1,"B14",True,"AD1457")
barchart("Top 12 por Indice H",5,13,4,"J14",True,"1F4E78")
barchart("Top 12 por Publicaciones",8,13,7,"B31",True,"2E7D32")

colh=BarChart(); colh.type="col"; colh.title="Distribucion de Indice H"; colh.legend=None; colh.height=7.5; colh.width=13
colh.add_data(Reference(dz,min_col=11,min_row=1,max_row=7),titles_from_data=True); colh.set_categories(Reference(dz,min_col=10,min_row=2,max_row=7))
try: colh.series[0].graphicalProperties.solidFill="EF6C00"
except: pass
db.add_chart(colh,"J31")

colc=BarChart(); colc.type="col"; colc.title="Distribucion por Citas"; colc.legend=None; colc.height=7.5; colc.width=13
colc.add_data(Reference(dz,min_col=14,min_row=1,max_row=7),titles_from_data=True); colc.set_categories(Reference(dz,min_col=13,min_row=2,max_row=7))
try: colc.series[0].graphicalProperties.solidFill="00838F"
except: pass
db.add_chart(colc,"B48")

pie=PieChart(); pie.title="Cobertura ORCID"; pie.height=7.5; pie.width=9
pie.add_data(Reference(dz,min_col=17,min_row=1,max_row=4),titles_from_data=True); pie.set_categories(Reference(dz,min_col=16,min_row=2,max_row=4))
db.add_chart(pie,"J48")

pie2=PieChart(); pie2.title="Estado de registros"; pie2.height=7.5; pie2.width=9
pie2.add_data(Reference(dz,min_col=20,min_row=1,max_row=4),titles_from_data=True); pie2.set_categories(Reference(dz,min_col=19,min_row=2,max_row=4))
db.add_chart(pie2,"P48")

# ---------- Metodologia ----------
mt=wb.create_sheet("Metodologia"); mt.column_dimensions["A"].width=110
N=[("BASE DE DATOS DE INVESTIGADORES UANL",True),("",False),
("Fecha: 28-jul-2026 | 93 investigadores unicos (102 filas con duplicados).",False),
("85 con perfil bibliometrico en OpenAlex; 70 con ORCID; 8 sin perfil (revisar en WoS).",False),("",False),
("COLUMNAS COMPLETAS (los 93): ORCID, URL, Total Publicaciones, Total Citas, Indice H, Indice i10, Citas Promedio 2yr, Citas por Pub, Estado.",False),
("COLUMNAS PARCIALES: tipos de trabajo y Revistas Principales completos para 10 (piloto); el resto 'Pendiente' (tarea programada tras reset de OpenAlex).",False),
("CUARTILES (Pub Q1-Q4): pendientes de JCR oficial (Web of Science, requiere login del usuario).",False),("",False),
("FUENTES: CVU/Nombre del archivo original; metricas bibliometricas de OpenAlex (api.openalex.org).",False),
("DESAMBIGUACION: emparejado por nombre completo + ORCID + afiliacion UANL. Ver columna Notas para avisos de homonimos/perfiles fragmentados.",False),
("DASHBOARD: KPIs y graficas en la pestaña 'Dashboard' (top por citas/H/publicaciones, distribuciones, cobertura ORCID, estado).",False),
]
for i,(txt,b) in enumerate(N,1):
    x=mt.cell(i,1,txt); x.font=Font(name="Arial",size=11,bold=b,color=NAVY if b else "000000"); x.alignment=Alignment("left","center",wrap_text=True)

wb.move_sheet("Dashboard",-(len(wb.sheetnames)-1))  # dashboard primero
wb.active=wb.sheetnames.index("Dashboard")
wb.save(OUT)
print("OK | uni:",len(uni),"pubs:",sumpubs,"cites:",sumcites,"Havg:",havg,"Hmax:",hmax)
